from openpyxl import load_workbook
file = 'D:/Code/cohort4/Python/Comp 230/datasheet.xlsx'


def read_datasheet():
    wb = load_workbook(file)
    d = {}  # new dictionary
    for ws in wb.sheetnames:
        d[ws] = {}  # add worksheets
        i = 1
        for row in wb[ws].iter_rows(min_row=3, min_col=1):
            if row[0].value is not None:
                d[ws][i] = {}  # add rows
                col = 0
                for cell in row:
                    d[ws][i][wb[ws][2][col].value] = cell.value
                    col += 1
            i += 1
    return(d)


read_datasheet()
