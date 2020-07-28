from openpyxl import load_workbook
import invoice_tracker
import json

file = 'D:/Code/cohort4/Python/Comp_230/invoice.xlsx'
d = invoice_tracker.read_datasheet()


def invoice():
    # ask user for invoice number
    inv_No = int(input('Please enter invoice #: '))

    # check if the invoice exists
    for rows in d['Invoices'].values():
        if rows['invoice_no'] == inv_No:
            found = True
            break
        else:
            found = False
    if not found:
        raise KeyError("Invoice not found")

    # get Customer & invoice info from the dictionary
    invoice = d['Invoices'][inv_No]
    customer = d['Customers'][invoice['customer_id']]

    # keys to print from customer and invoice
    k_customer = ('customer_id', 'f_name', 'l_name', 'phone')
    k_invoice = ('invoice_date', 'payment_method', 'total_price')

    # create a new worksheet
    wb = load_workbook(file)
    ws = wb.create_sheet()
    ws.title = f"Invoice {inv_No}"
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.cell(1, 1).value = f'INVOICE #{inv_No}'

    # add the info to the worksheet
    row = 2
    col = 2
    for key in k_customer:
        ws.cell(row, col).value = json.dumps(key)[1:-1]
        ws.cell(row, col+1).value = json.dumps(customer[key])
        row += 1

    for key in k_invoice:
        ws.cell(row, col).value = json.dumps(key)[1:-1]
        ws.cell(row, col+1).value = json.dumps(invoice[key])
        row += 1

    wb.save(file)
    print(f'Printed inv_No {inv_No}')


invoice()
